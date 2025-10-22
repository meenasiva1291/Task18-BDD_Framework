import json
import openpyxl
from behave import given, when, then
from behave.exception import StepNotImplementedError
from selenium import webdriver
from selenium.common import NoAlertPresentException, TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.expected_conditions import alert_is_present
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC, wait

#Open browser and load URL
@given(u'I open zenportal login page')
def step_impl(context):
        context.driver = webdriver.Chrome()
        context.driver.get("https://www.zenclass.in/login")

#Method to read data from excel
def read_credentials_from_excel(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    for i in range(2, sheet.max_row + 1):
        email = sheet.cell(row=i, column=1).value
        password = sheet.cell(row=i, column=2).value
        data.append((email, password))
    return data

#Take username and password from Excel
@when(u'I login using credentials from Excel file')
def step_impl(context):
    data = read_credentials_from_excel("C://Users//HP//PycharmProjects//Task18-BDD_Framework//testdata//test_data.xlsx", "Sheet1")
    for email, password in data:
        context.driver.find_element(By.XPATH,"//input[@placeholder='Enter your mail']").clear()
        context.driver.find_element(By.XPATH,"//input[@placeholder='Enter your password ']").clear()
        context.driver.find_element(By.XPATH,"//input[@placeholder='Enter your mail']").send_keys(email)
        context.driver.find_element(By.XPATH,"//input[@placeholder='Enter your password ']").send_keys(password)

#Take username and password from Excel
@when(u'I login using credentials from json file')
def step_login_json(context):
    with open("C://Users//HP//PycharmProjects//Task18-BDD_Framework//testdata//jsontest_data.json", "r") as file:
        data = json.load(file)
    user = data["user"][0]  # first user only
    email = user["email"]
    password = user["password"]

    context.driver.find_element(By.XPATH, "//input[@placeholder='Enter your mail']").clear()
    context.driver.find_element(By.XPATH, "//input[@placeholder='Enter your password ']").clear()
    context.driver.find_element(By.XPATH, "//input[@placeholder='Enter your mail']").send_keys(email)
    context.driver.find_element(By.XPATH, "//input[@placeholder='Enter your password ']").send_keys(password)

#Enter valid email and password
@when(u'I enter valid username and password')
def step_impl(context):
        context.driver.find_element(By.XPATH,"//input[@placeholder='Enter your mail']").send_keys("meenakshisivasankar@gmail.com")
        context.driver.find_element(By.XPATH,"//input[@placeholder='Enter your password ']").send_keys("Meenu1291")

#validate login button
@when(u'I click on submit button')
def step_impl(context):
        context.driver.find_element(By.XPATH, "//button[@type='submit']").click()

#validate landing page on successful login
@then(u'I should be directed to dashboard page')
def step_impl(context):
    try:
        WebDriverWait(context.driver, 2).until(EC.alert_is_present())
        alert = context.driver.switch_to.alert
        print("Alert message:", alert.text)
        alert.accept()
        print("Alert closed successfully.")

    except (TimeoutException, NoAlertPresentException):
        print("No alert appeared after clicking submit — proceeding to dashboard check.")

    # Wait for dashboard URL or dashboard element
    wait = WebDriverWait(context.driver, 10)
    wait.until(EC.url_contains("dashboard"))

    actual_url = context.driver.current_url
    expected_url = "https://www.zenclass.in/dashboard"

    # Assert that we actually reached dashboard
    assert expected_url in actual_url, {f"URL mismatch: expected {expected_url}, got {actual_url}"}

    print("Successfully logged into the dashboard page:", actual_url)

#validate invalid username and password
@when(u'I enter invalid username and password')
def step_impl(context):
    context.driver.find_element(By.XPATH, "//input[@placeholder='Enter your mail']").send_keys("abc@gmail.com")
    context.driver.find_element(By.XPATH, "//input[@placeholder='Enter your password ']").send_keys("Mass111")

#validate error message on entering invalid username and password
@then(u'I should receive an error message')
def step_impl(context):
    wait = WebDriverWait(context.driver, 10)

    try:
        # Correct tuple format passed to visibility_of_element_located
        error_msg = wait.until(
            EC.visibility_of_element_located((By.XPATH, "//p[text()='Invalid email!']"))
        )

        # Extract and verify error text
        actual_text = error_msg.text.strip()
        expected_text = "Invalid email!"

        assert actual_text == expected_text, f" Expected '{expected_text}', but got '{actual_text}'"
        print("Error message displayed correctly:", actual_text)

    except TimeoutException:
        print("Error message not displayed within 10 seconds.")
        raise

#validate input box
@then(u'username and password input box is visible and enabled')
def step_impl(context):
    wait = WebDriverWait(context.driver, 15)

    try:
        # Wait for username field
        username_box = wait.until(
            EC.visibility_of_element_located((By.XPATH, "//input[contains(@placeholder, 'mail') or contains(@placeholder, 'email')]"))
        )
        print("Username box is visible and enabled.")

        # Wait for password field (using more flexible locator)
        password_box = wait.until(
            EC.visibility_of_element_located((By.XPATH, "//input[contains(@placeholder, 'password') or @type='password']"))
        )
        print("Password box is visible and enabled.")

        assert username_box.is_enabled() and password_box.is_enabled(), "Input boxes not enabled."

    except TimeoutException:
        print("Timeout: Username or password box not found or not visible within wait time.")
        # Optional debug info
        print("Current page URL:", context.driver.current_url)
        print("Page title:", context.driver.title)
        raise

    except NoSuchElementException:
        print("Locator issue: Element not found on the page.")
        raise

#validate logout functionality
@then(u'I should be able to click on logout')
def step_impl(context):
    context.driver.find_element(By.XPATH,"//button[@class='custom-close-button']").click()
    context.driver.find_element(By.XPATH,'//div[@class="profile-click-icon-div"]').click()
    context.driver.find_element(By.XPATH, "//div[contains(@class,'notify-container')]/div[3]").click()
    print("Logout button clicked.")
